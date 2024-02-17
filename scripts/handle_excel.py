from openpyxl import load_workbook
from scripts.handle_config import do_config


class HandleExcel:
    """
    处理Excel文件中测试数据的操作类
    """
    def __init__(self,filename,sheet_name=None):
        """
        创建属性，并初始化
        :param filename: excel文件名
        :param sheet: sheet表单名为默认参数，不传参数时，默认获取第一个激活的表单，传参时使用传递sheet表单名
        """
        self.filename = filename
        self.sheet_name = sheet_name

    def get_cases(self):
        """
        获取文件所有的测试用例数据
        :return: 返回嵌套字典的列表
        """
        wb = load_workbook(self.filename)
        if self.sheet_name is None:
            ws = wb.active
        else:
            ws = wb[self.sheet_name]
        head_data_tuple = tuple(ws.iter_rows(max_row=1,values_only=True))[0]
        one_list = []
        for one_tuple in ws.iter_rows(min_row=2,values_only=True):
            one_list.append(dict(zip(head_data_tuple,one_tuple)))   # 将表头所在行的元组数据作为key，其他数据行的元组作为value
        wb.close()
        return one_list

    def get_one_case(self,row):
        """
        获取第几条测试用例数据
        :param row:
        :return:
        """
        return  self.get_cases()[row-1]

    def write_result(self,row,actual,result):
        """
         在指定行写入数据
        :param row: 写入第几行
        :param actual: 服务器返回的实际结果
        :param result: 测试用例的执行结果
        :return:
        """
        other_wb = load_workbook(self.filename)
        if self.sheet_name is None:
            ws = other_wb.active
        else:
            ws = other_wb[self.sheet_name]
        if isinstance(row,int) and 2 <= row <=ws.max_row:
            ws.cell(row=row,column=do_config.get_int("excel","actual_column"),value=actual)
            ws.cell(row=row,column=do_config.get_int("excel","result_column"),value=result)
            other_wb.save(self.filename)
            print(f"第{row}行写入测试数据actual：【{actual}】，result：【{result}】成功！")
        else:
            print("输入行号有误，应大于1的整数！")
        other_wb.close()

if __name__ == '__main__':
    do_excel = HandleExcel(r"C:\Users\78228\Desktop\automated_testing_api\TestApi\cases.xlsx")
    print(do_excel.get_cases())
    # do_excel.write_result(2,10,"pass")





